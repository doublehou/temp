#!/usr/bin/env python

from openpyxl import Workbook
from openpyxl import load_workbook

file='a.xlsx'
file2='libreoffice-saved.xlsx'
workbook=load_workbook(file)
workbook2=load_workbook(file2)
sheet=workbook.get_active_sheet()
sheet2=workbook2.get_active_sheet()

sheet['A1']='somestring' #modify 
sheet2['A1']='somestring'
workbook.save(filename='modify.xlsx')
workbook2.save(filename='libreoffice-modify.xlsx')
